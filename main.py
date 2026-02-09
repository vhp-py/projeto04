#importando as bibliotecas necessárias
import json
from pathlib import Path
from openpyxl import Workbook, load_workbook




caminho_arquivo = Path('dados_loja.json')
dados_loja = {
    'produtos': [], 
    'caixa': 0.0,
}

# --- 2. CARREGAR DADOS ---
if caminho_arquivo.exists():
    try:
        if caminho_arquivo.stat().st_size > 0:
            with open(caminho_arquivo, mode='r', encoding='utf-8') as file:
                dados_loja = json.load(file)
        else:
            with open(caminho_arquivo, mode='w', encoding='utf-8') as file:
                json.dump(dados_loja, file, indent=4)
    except Exception as e:
        print(f'Erro ao carregar: {e}')
else: 
    with open(caminho_arquivo, mode='w', encoding='utf-8') as file:
        json.dump(dados_loja, file, indent=4, ensure_ascii=False)


# ---  FUNÇÕES ---
         
def comprar_produto():
    produto_escolhido = input('Digite o produto que deseja pesquisar: ')
    encontrou = False
    
       
    lista_de_produtos = dados_loja['produtos']
    
    for item in lista_de_produtos:
         
        try:
            if produto_escolhido == item['nome_produto']:
                print(f"Achei: {item['nome_produto']}")
                print(f"Preço unitario: R${item['preco_de_venda']}")
                qtd_comprar = int(input('Quantos quer levar? '))
                
                if qtd_comprar > item['quantidade']:
                    print(f'Possuimo apenas {item['quantidade']} de {item['nome_produto']} no estoque.')
                    break
                
                
                preco_total = qtd_comprar * item['preco_de_venda']
                print(f'Levando {qtd_comprar} {item['nome_produto']}, fica R${preco_total:.2f}')

                finalizar_compra = input('Finalizar compra ? y/n ').lower()
                if finalizar_compra == 'y':
                    dados_loja['caixa'] += preco_total
                    item['quantidade'] -= qtd_comprar
                    salvar_dados()
                    print('Compra finalizada.\nObrigado pela preferência!')
                    
                elif finalizar_compra == 'n':
                    print('Saindo...')
                    break
        except ValueError:
            print(f'Erro: Digite apenas numero')

def cadastrar_produto():
    print('='*50)
    print("\nNOVO CADASTRO\n")
    print('='*50)
    try:
         
        nome = input('Nome do Produto: ')
        qtd = int(input('Digite a quantidade: '))
        custo = float(input('Digite o preço de custo: '))
        venda = float(input('Digite o preço de venda: '))

        
        novo_produto = {
            'nome_produto': nome,
            'quantidade': qtd,
            'preco_de_custo': custo,
            'preco_de_venda': venda
        }

        
        dados_loja['produtos'].append(novo_produto)
        salvar_dados() 
        print(f"Produto '{nome}' cadastrado com sucesso!")

    except ValueError:
        print("Erro: Você digitou letras em campos de números!")

def listar_produtos():
    print("\n--- LISTA DE PRODUTOS ---")
   
    lista = dados_loja['produtos'] 
    
    for item in lista:
        if item['quantidade'] == 0:
            print(f'Estamos sem estoque de {item['nome_produto']}')
            
        print(f"Produto: {item['nome_produto']} | Preço: R$ {item['preco_de_venda']}")

def salvar_dados():
    """Função auxiliar para salvar sempre que mudarmos algo"""
    with open(caminho_arquivo, mode='w', encoding='utf-8') as file:
        json.dump(dados_loja, file, indent=4, ensure_ascii=False)

def exportar_dados():
    arquivo_exportado = Workbook()
    #   Aba de estoque
    aba_estoque = arquivo_exportado.active
    aba_estoque.title = 'Estoque'

    aba_estoque.append(['Produto', 'Quantidade', 'Custo Un.', 'Total Investido'])
    
    for item in dados_loja['produtos']:
        total_investido = item['quantidade'] * item['preco_de_custo']
        
        aba_estoque.append([
            item['nome_produto'],
            item['quantidade'],
            item['preco_de_custo'],
            total_investido
        ])
    
    #   Aba de resumos
    aba_resumo = arquivo_exportado.create_sheet('Resumo Financeiro')
    
    aba_resumo.append(['Resumo', 'Valor Total'])
    aba_resumo.append(['Dinheiro em caixa', dados_loja['caixa']])

    for celula in aba_estoque['D']:
        if celula.row > 1:
            celula.number_format = 'R$ #,##0.00' 

    arquivo_exportado.save('relatorio_financeiro.xlsx')
    print('Relatório Criado com Sucesso')

    



def main():
    while True: # Loop infinito para o menu não fechar
        print('\n STORE MANAGER ')
        print('[1] - Fazer compra')
        print('[2] - Gerenciar Estoque')
        print('[3] - Sair')
        
        opcao = input('Digite a opção desejada: ')

        if opcao == '1':
            comprar_produto()
        
        elif opcao == '2':
            # Sub-menu de Estoque
            print('\n-- ESTOQUE --')
            print('[1] - Cadastrar Produto')
            print('[2] - Listar Produtos')
            print('[3] - Fechar caixa e Exportar')
            op_estoque = input('Escolha: ')

            if op_estoque == '1':
                cadastrar_produto()
            elif op_estoque == '2':
                listar_produtos()
            elif op_estoque == '3':
                exportar_dados()
        elif opcao == '3':
            print("Saindo do sistema...")
            break
        
        else:
            print("Opção inválida!")

# Executa o programa
main()
